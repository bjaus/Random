import os, csv, sys

from openpyxl import Workbook
from openpyxl.cell import get_column_letter

my_path = r'C:\Users\BJaus\Desktop\TEST'

csv_filename = sys.argv[1]
csv_filepath = my_path 
csv_filepath = csv_filepath + '\\' + csv_filename

f = open(csv_filepath, 'rU')

csv.register_dialect('comma', delimiter=',')

reader = csv.reader(f, dialect='comma')

wb = Workbook()
dest_filename = sys.argv[2]
dest_filepath = my_path + '\\' + dest_filename

ws = wb.worksheets[0]
ws.title = 'TEST'

for row_index, row in enumerate(reader):
	for column_index, cell in enumerate(row):
		column_letter = get_column_letter((column_index + 1))
		ws.cell('%s%s'%(column_letter, (row_index + 1))).value = cell

wb.save(filename = dest_filename)